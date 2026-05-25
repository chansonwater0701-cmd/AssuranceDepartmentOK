#!/usr/bin/env python
# -*- coding: utf-8 -*-
import sys
import os

# Ensure openpyxl is available
try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl
    from openpyxl import load_workbook

# Load the Excel file
file_path = r"C:\Users\rpa\Downloads\千山機台維修建議(20260522).xlsx"

try:
    print("=" * 100)
    print(f"Reading Excel File: {file_path}")
    print("=" * 100)
    
    wb = load_workbook(file_path)
    
    # Get all sheet names
    sheet_names = wb.sheetnames
    print(f"\nTotal number of sheets: {len(sheet_names)}")
    print(f"Sheet names: {sheet_names}\n")
    
    # Iterate through each sheet
    for sheet_idx, sheet_name in enumerate(sheet_names, 1):
        ws = wb[sheet_name]
        print("\n" + "=" * 100)
        print(f"SHEET {sheet_idx}: {sheet_name}")
        print("=" * 100)
        
        # Get dimensions
        max_row = ws.max_row
        max_col = ws.max_column
        
        print(f"Dimensions: {max_row} rows x {max_col} columns\n")
        print("-" * 100)
        
        # Read all data from the sheet
        for row_idx in range(1, max_row + 1):
            row_data = []
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_value = cell.value
                # Format: Column Header (or A, B, C...): value
                col_letter = cell.column_letter
                row_data.append(f"{col_letter}{row_idx}: {repr(cell_value)}")
            
            print(f"Row {row_idx:>4d} | {' | '.join(row_data)}")
        
        print("-" * 100)
    
    wb.close()
    print("\n" + "=" * 100)
    print("Excel file reading completed successfully!")
    print("=" * 100)

except FileNotFoundError:
    print(f"ERROR: File not found at: {file_path}")
    print(f"Please check if the file exists.")
    sys.exit(1)
except Exception as e:
    print(f"ERROR: {str(e)}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
