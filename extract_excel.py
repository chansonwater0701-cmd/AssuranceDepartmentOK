#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
import sys
import os

excel_file = r"C:\Users\rpa\Downloads\千山機台維修建議(20260522).xlsx"

if not os.path.exists(excel_file):
    print(f"ERROR: File not found: {excel_file}")
    sys.exit(1)

try:
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    
    print("\n" + "="*120)
    print(f"EXCEL FILE: {excel_file}")
    print("="*120)
    
    sheet_names = wb.sheetnames
    print(f"\nTotal Sheets: {len(sheet_names)}")
    print(f"Sheet Names: {sheet_names}\n")
    
    for sheet_idx, sheet_name in enumerate(sheet_names, 1):
        ws = wb[sheet_name]
        max_row = ws.max_row
        max_col = ws.max_column
        
        print("\n" + "="*120)
        print(f"SHEET {sheet_idx}: [{sheet_name}]")
        print("="*120)
        print(f"Dimensions: {max_row} rows × {max_col} columns")
        print("-"*120)
        
        # Print all cells
        for row_idx in range(1, max_row + 1):
            row_cells = []
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_value = cell.value
                col_letter = cell.column_letter
                
                # Format cell reference and value
                if cell_value is None:
                    row_cells.append(f"{col_letter}: <EMPTY>")
                else:
                    # Clean up representation
                    row_cells.append(f"{col_letter}: {cell_value}")
            
            print(f"Row {row_idx:3d} | {' | '.join(row_cells)}")
        
        print("-"*120)
    
    wb.close()
    print("\n" + "="*120)
    print("Successfully completed!")
    print("="*120 + "\n")

except Exception as e:
    print(f"ERROR: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
