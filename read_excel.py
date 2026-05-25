import openpyxl
from openpyxl import load_workbook

file_path = r"C:\Users\rpa\Downloads\千山機台維修建議(20260522).xlsx"
wb = load_workbook(file_path)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"SHEET: {sheet_name}")
    print(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns")
    
    for row_idx in range(1, ws.max_row + 1):
        row_values = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            row_values.append(f"Col{col_idx}: {repr(cell.value)}")
        print(f"Row {row_idx}: " + " | ".join(row_values))
    
    print()

wb.close()
